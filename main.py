import time
import requests
import random
from bs4 import BeautifulSoup
from openpyxl import Workbook

hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num = 0
    book_list = []

    while(True):
        url = "http://book.douban.com/tag/{}?start={}&type={}".format(
            book_tag, page_num*20, 'T')
        print(url)
        plain_text = ""
        try:
            r = requests.get(url, headers=hds[page_num % len(hds)])
            plain_text = r.text
        except Exception as e:
            print(e)

        soup = BeautifulSoup(plain_text, "html.parser")
        books = soup.find_all('li', {'class', 'subject-item'})
        print(book_tag, page_num, len(books))
        if len(books) == 0:
            break
        for b in books:
            name, rating, people_num, info = get_book_info(b)
            print(name, rating, people_num, info)
            book_list.append(
                [name, rating, people_num, info])
            time.sleep(random.random()*5)
        page_num += 1
        print('Downloading Information From Page %d' % page_num)
        time.sleep(random.random()*5)
    return book_list


def get_book_info(dom):
    name = dom.find('div', {'class': 'info'}).find('h2').find('a').get_text()
    info = dom.find('div', {'class': 'info'}).find(
        'div', {'class': 'pub'}).get_text()
    rating = dom.find('div', {'class': 'info'}).find(
        'span', {'class': 'rating_nums'}).get_text()
    people_num = dom.find('div', {'class': 'info'}).find(
        'span', {'class': 'pl'}).get_text()
    return [name, rating, people_num, info]


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook()
    ws = []
    for i in range(len(book_tag_lists)):
        # utf8->unicode
        ws.append(wb.create_sheet(title=book_tag_lists[i]))
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号', '书名', '评分', '评价人数', '信息'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], bl[1],
                          bl[2], bl[3]])
            count += 1
    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-'+book_tag_lists[i])
    save_path += '.xlsx'
    wb.save(save_path)


if __name__ == '__main__':
    # book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    # book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    # book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    # book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    # book_tag_lists = ['数学']
    # book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    # book_tag_lists = ['商业','理财','管理']
    book_tag_lists = ['经济学', '管理', '经济', '商业', '金融',
                      '投资', '营销', '理财', '创业', '股票', '广告', '企业史', '策划']
    # book_tag_lists = ['科普','经典','生活','心灵','文学']
    # book_tag_lists = ['科幻','思维','金融']
    # book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
    book_lists = do_spider(book_tag_lists)
    # print(book_lists)
    print_book_lists_excel(book_lists, book_tag_lists)
    # print(get_book_info('https://book.douban.com/subject/1449351/'))
